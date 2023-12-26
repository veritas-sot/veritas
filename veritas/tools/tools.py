import colorlog
import yaml
import os
import getpass
import pytricia
import hashlib
import sys
from loguru import logger
from openpyxl import load_workbook
from loguru import logger
from veritas.messagebus import messagebus


def get_miniapp_config(appname, app_path, config_file=None):
    """return config of miniapp
    
    1. prio: config in local directory
    2. prio: config in local ./conf/ directory
    3. prio: config in homedirectory ~/.veritas
    4. prio: config in /etc/veritas/
    """

    config_filename = config_file if config_file else f'{appname}.yaml'
    local_config_file = f'{app_path}/{config_filename}'
    local_subdir_config_file = f'{app_path}/conf/{config_filename}'
    homedir_config_file = f'{os.path.expanduser("~")}/.veritas/miniapps/{appname}/{config_filename}'
    etc_config_file = f'/etc/veritas/miniapps/{appname}/{config_filename}'

    if os.path.exists(local_config_file):
        filename = local_config_file
    elif os.path.exists(local_subdir_config_file):
        filename = local_subdir_config_file
    elif os.path.exists(homedir_config_file):
        filename = homedir_config_file
    elif os.path.exists(etc_config_file):
        filename = etc_config_file
    else:
        logger.critical(f'neither {local_config_file} nor {local_subdir_config_file} ' \
                        f'or {etc_config_file} exist')
        return None

    logger.debug(f'reading {filename}')
    with open(filename) as f:
        try:
            return yaml.safe_load(f.read())
        except Exception as exc:
            logger.error(f'could not read or parse config')
            return None

def create_logger_environment(config, cfg_loglevel=None, cfg_loghandler=None):
    """return database, zeromq and formatter"""

    loglevel = cfg_loglevel.upper() if cfg_loglevel \
        else config.get('general',{}).get('logger',{}).get('loglevel', 'INFO')
    handler_txt = cfg_loghandler if cfg_loghandler \
        else config.get('general',{}).get('logger',{}).get('handler', 'sys.stdout')
    
    # evaluate handler
    if handler_txt == 'sys.stdout' or handler_txt == 'stdout':
        loghandler = sys.stdout
    elif handler_txt == 'sys.stderr' or handler_txt == 'stderr':
        loghandler = sys.stderr
    else:
        loghandler = handler_txt

    if config.get('general',{}).get('logger',{}).get('logtodatabase', False):
        database = config.get('general',{}).get('logger',{}).get('database')
    else:
        database = None

    if config.get('general',{}).get('logger',{}).get('logtozeromq', False):
        zeromq = config.get('general',{}).get('logger',{}).get('zeromq')
    else:
        zeromq = None

    # configure logger
    if loglevel.upper() == "DEBUG":
        logger_format = (
                "<green>{time:YYYY-MM-DD HH:mm:ss.SSS}</green> | "
                "<level>{level: <8}</level> | "
                "<cyan>{name}</cyan>:<cyan>{function}</cyan>:<cyan>{line}</cyan> | "
                "{extra[extra]} | <level>{message}</level>"
        )
    else:
        logger_format = (
                "<green>{time:YYYY-MM-DD HH:mm:ss.SSS}</green> | "
                "<level>{level: <8}</level> | "
                "{extra[extra]} | <level>{message}</level>"
        )

    logger.configure(extra={"extra": "unset"})
    logger.remove()
    logger.add(loghandler, level=loglevel, format=logger_format)
    if database or zeromq:
        logger.debug(f'enabling veritas messagebus db: {database != None} zeroMQ: {zeromq != None}')
        logger.add(messagebus.Messagebus(database=database,
                                         use_queue=False,
                                         zeromq=zeromq,
                                         app='onboarding'),
                level=loglevel)

    return database, zeromq, loglevel, loghandler, logger_format

def get_value_from_dict(dictionary, keys):
    if dictionary is None:
        return None

    nested_dict = dictionary

    for key in keys:
        try:
            nested_dict = nested_dict[key]
        except KeyError as e:
            return None
        except IndexError as e:
            return None
        except TypeError as e:
            return nested_dict

    return nested_dict

def convert_arguments_to_properties(*unnamed, **named):
    """ convert unnamed (dict) and named arguments to a single property dict """
    properties = {}
    if len(unnamed) > 0:
        for param in unnamed:
            if isinstance(param, dict):
                for key,value in param.items():
                    properties[key] = value
            elif isinstance(param, str):
                # it is just a text like log('something to log')
                return param
            elif isinstance(param, tuple):
                for tup in param:
                    if isinstance(tup, dict):
                        for key,value in tup.items():
                            properties[key] = value
                    elif isinstance(tup, str):
                        return tup
                    elif isinstance(tup, list):
                        return tup
            elif isinstance(param, list):
                return param
            else:
                logger.error(f'cannot use paramater {param} / {type(param)} as value')
    for key,value in named.items():
            properties[key] = value
    
    return properties

def get_username_and_password(args, sot, config):
    username = None
    password = None

    if args.profile is not None:
        username = config.get('profiles',{}).get(args.profile,{}).get('username')
        token = config.get('profiles',{}).get(args.profile,{}).get('password')
        if username and token:
            auth = sot.auth(encryption_key=os.getenv('ENCRYPTIONKEY'), 
                            salt=os.getenv('SALT'), 
                            iterations=int(os.getenv('ITERATIONS')))
            password = auth.decrypt(token)

    # overwrite username and password if configured by user
    username = args.username if args.username else username
    password = args.password if args.password else password

    username = input("Username (%s): " % getpass.getuser()) if not username else username
    password = getpass.getpass(prompt="Enter password for %s: " % username) if not password else password

    return username, password

def read_excel_file(filename):

    response = {}
    table = []

    # Load the workbook
    workbook = load_workbook(filename = filename)

    # Select the active worksheet
    worksheet = workbook.active
    
    # loop through table and build list of dict
    rows = worksheet.max_row
    columns = worksheet.max_column + 1 
    for row in range(2, rows + 1):
        line = {}
        for col in range(1, columns):
            key = worksheet.cell(row=1, column=col).value
            value = worksheet.cell(row=row, column=col).value
            line[key] = value
        table.append(line)

    return table

def set_value(mydict, paths, value):
    # write value to nested dict
    # we split the path by using '__'
    parts = paths.split('__')
    for part in parts[0:-1]:
        # add {} if item does not exists
        # this loop create an empty path
        mydict = mydict.setdefault(part, {})
    # at last write value to dict
    mydict[parts[-1]] = value

def get_prefix_path(prefixe, ip):
    """return prefix path"""
    prefix_path = []
    pyt = pytricia.PyTricia()

    # build pytricia tree
    for prefix_ip in prefixe:
        pyt.insert(prefix_ip, prefix_ip)

    try:
        prefix = pyt.get(ip)
    except Exception as exc:
        logger.error(f'prefix not found; using 0.0.0.0/0')
        prefix = "0.0.0.0/0"
    prefix_path.append(prefix)

    parent = pyt.parent(prefix)
    while (parent):
        prefix_path.append(parent)
        parent = pyt.parent(parent)
    return prefix_path[::-1]

def calculate_md5(row):
    data = ""
    for d in row:
        if isinstance(d, list):
            my_list = ''.join(d)
            data += my_list
        elif d is None or d == 'null':
            pass
        else:
            data += d
    return hashlib.md5(data.encode('utf-8')).hexdigest()

def flatten_json(y):
    """flatten json"""
    # source https://towardsdatascience.com/flattening-json-objects-in-python-f5343c794b10
    out = {}

    def flatten(x, name=''):
        if type(x) is dict:
            for a in x:
                flatten(x[a], name + a + '_')
        elif type(x) is list:
            i = 0
            for a in x:
                flatten(a, name + str(i) + '_')
                i += 1
        else:
            out[name[:-1]] = x

    flatten(y)
    return out
